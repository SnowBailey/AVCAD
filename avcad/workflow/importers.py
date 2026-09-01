"""清单导入富集层：xlsx / 自由格式清单 -> 规范化引擎条目。

职责（清单驱动 5 步工作流的第①步「AI 解析品牌/型号/数量/类别」）：
- 吊架/飞行架/桁架等非信号设备：关键词命中即排除，优先级高于主库（避免主库误判为扬声器）。
- 品牌+型号优先走 product_resolver 主库（eko_catalog）补全类别/特性/参数/电气。
- 主库未覆盖或延迟（deferred，如 QU-16）的型号：用设备名关键词兜底分类 + 指标参数正则抽取。
- 功放归一化：主库给 channels_hint，需映射为 channels，并从指标参数补 power_w_per_ch 电气功率。

纯标准库 + openpyxl（已是 avcad 依赖）。不引入新依赖。
"""
from __future__ import annotations
import csv
import io
import json
import re

from avcad.parse.bom_parser import dump_params
from avcad.parse.product_resolver import enrich as resolve_products

# 非信号设备：机械/结构件，不出系统图
RIGGING = ["吊架", "飞行架", "桁架"]
# 电源类辅助设备：电源时序器、电源管理器、电源控制器等。
# 它们是配电系统的一部分，但不在音频信号链路上——出图会让画图逻辑误判
# （友腾 EAW2/3 清单里有 2 台「电源时序器」被兜底成 IO，既错又孤立）。
# 这里走 drop 路径，与 RIGGING 同处理（is_rigging 也复用此判断）。
POWER_AUX = ["电源时序器", "电源管理器", "电源控制器", "电源监测"]
# 占位项：造价清单里的「甲方自配 / 按需保留 / 待定」等，无品牌无型号，
# 无法建模，也不该让用户每次在第②步手动点「不需要」。
# ★ 只在 brand 与 model 同时为空时才排除，避免误杀真实设备。
PLACEHOLDER = ["按需保留", "自配", "甲方自配", "甲供", "待定", "暂定",
               "按实结算", "详见图纸", "另计", "不含", "见清单"]
# 成对销售：清单「单位=对/副/pair」时按 2 台/支展开
# （IPS UM2000AP 全指向天线、UM2000AT 有源指向天线：主库 remark「1 对 = 2 支」）
PAIR_UNITS = {"对", "副", "pair", "pairs"}
PAIR_UNIT_FACTOR = 2
# 设备名关键词 -> 类别（兜底分类，顺序即优先级）
CATEGORY_KW = [
    ("SOURCE", ["话筒", "麦克风", "传声器", "音源"]),
    ("WIRELESS_MIC", ["发射", "无线话筒"]),
    ("WIRELESS_RX", ["接收机", "接收端"]),
    ("ANTENNA", ["天线"]),
    ("IO", ["接口箱", "接口卡", "接口矩阵"]),
    ("SWITCH", ["交换机"]),
    # 「调音」单独成词：真实清单有「数字调音控制盘」「数字调音控制台」等写法，
    # 只匹配「调音台」会漏（A&H QU-16 在菏泽清单里就叫「数字调音控制盘」）。
    ("MIXER", ["调音台", "调音", "控制台", "混音"]),
    ("PROCESSOR", ["处理器", "dsp", "音频处理"]),
    ("AMP", ["功率放大器", "功放", "放大器"]),
    ("SPEAKER", ["扬声器", "音箱", "全频", "线阵列", "低音", "补声", "返送",
                 "返听", "主扩", "拉声像", "台唇", "环绕", "音柱", "扩声", "号筒"]),
]
# 控制/网络特性关键词
CONTROL_KW = ["集控", "网络监测", "网络检测", "tcp/ip", "rs-232", "rs-485",
              "gpio", "控制", "状态上报", "ip控制", "中控"]

_AMP_CH = re.compile(r"(\d+)\s*通道")
_AMP_POWER = re.compile(r"8Ω.*?(\d+)\s*W", re.I | re.S)
_OHM = re.compile(r"(\d+)\s*Ω")
_SP_POWER = re.compile(r"(\d+)\s*W")
_IN = re.compile(r"模拟输入\s*(\d+)\s*路")
_OUT = re.compile(r"模拟输出\s*(\d+)\s*路")


def is_rigging(name: str) -> bool:
    """非信号设备判定：机械/结构件 + 电源类辅助设备（不出系统图）。

    阳哥 2026-08-31：「电源时序器」等配电设备不属于音频信号链路，
    应在第②步被排除出图，而不是被兜底成 IO 设备后孤立。
    """
    return any(k in (name or "") for k in RIGGING + POWER_AUX)


def is_placeholder(name: str, row=None) -> bool:
    """占位项判定：无品牌 + 无型号 + 名称是「自配 / 按需保留」之类。

    仅在品牌与型号同时为空时生效，否则「定制机柜」这类真实设备会被误杀。
    """
    if not any(k in (name or "") for k in PLACEHOLDER):
        return False
    row = row or {}
    brand = str(row.get("品牌") or row.get("brand") or "").strip()
    model = str(row.get("型号") or row.get("model") or "").strip()
    return not brand and not model


def classify_category(name: str):
    """返回类别或 None（吊架）。"""
    if is_rigging(name):
        return None
    for cat, kws in CATEGORY_KW:
        if any(k in (name or "") for k in kws):
            return cat
    return None


def extract_features(spec: str, name: str, category: str) -> set:
    feats = set()
    s = (spec or "").lower()
    n = (name or "").lower()
    if "dante" in s:
        feats.add("dante")
    if any(k in s or k in n for k in CONTROL_KW):
        feats.add("control")
    if category in ("PROCESSOR",) and ("模拟" in s or "analog" in s):
        feats.add("analog")
    if "有源" in (name or ""):
        feats.add("active")
    return feats


def extract_params(spec: str, category: str) -> dict:
    p = {}
    if not spec:
        return p
    if category == "AMP":
        m = _AMP_CH.search(spec)
        if m:
            p["channels"] = int(m.group(1))
        m = _AMP_POWER.search(spec)
        if m:
            p["power_w_per_ch"] = int(m.group(1))
    elif category == "SPEAKER":
        m = _OHM.search(spec)
        if m:
            p["impedance_ohm"] = int(m.group(1))
        m = _SP_POWER.search(spec)
        if m:
            p["power_w"] = int(m.group(1))
    elif category in ("PROCESSOR", "MIXER"):
        m = _IN.search(spec)
        if m:
            p["inputs"] = int(m.group(1))
        m = _OUT.search(spec)
        if m:
            p["outputs"] = int(m.group(1))
    return p


def _qty(v) -> int:
    try:
        return int(float(v or 1))
    except (ValueError, TypeError):
        return 1


# 表头别名 -> 标准列名（真实清单千差万别，大小写/换行/中英混用都要兜住）
HEADER_ALIASES = {
    "设备名称": ["设备名称", "产品名称", "名称", "设备名",
                "设备名称\nname of equipment", "name of equipment", "name"],
    "品牌": ["品牌", "品牌/国别", "品牌/产地", "厂商", "brand", "manufacturer",
            "品牌/国别\nbrand/country", "brand/country"],
    "型号": ["型号", "规格型号", "产品型号", "model", "model number",
            "型号\nmodel number"],
    "数量": ["数量", "台数", "qty", "quantity", "数量\nquantity"],
    "单位": ["单位", "计量单位", "unit"],
    "指标参数": ["指标参数", "产品参数", "规格", "规格参数", "技术参数", "参数",
                "index parameter", "spec"],
}
# 价目表（产品总表）判定阈值
# 真实报价文件常把「厂商全系价目表」挂在最后一个 sheet，几百行、数量列几乎全空/零
PRICELIST_MAX_COLS = 30      # 列数异常多 -> 价目表（EAW 配单表 2575 列）
PRICELIST_MIN_ROWS = 20      # 行数达到这个量才启用「零占比」判据
PRICELIST_ZERO_RATIO = 0.6   # 数量列空/零占比超过此值 -> 价目表


def _find_header_row(rows) -> int:
    """在前 12 行内找表头行：必须同时含「设备名称类」与「型号类」列。"""
    name_keys = set(HEADER_ALIASES["设备名称"])
    model_keys = set(HEADER_ALIASES["型号"])
    for i, r in enumerate(rows[:12]):
        cells = {str(c).strip().lower() for c in r if c is not None}
        if cells & name_keys and cells & model_keys:
            return i
    return -1


def _clean_brand(v) -> str:
    """品牌归一化：真实清单常写成「IPS/深圳」「Allen & Heath/英国」「ezacoustics/深圳」。

    主库里存的是纯品牌名（IPS / ALLEN&HEATH / ezacoustics），带国别后缀会
    导致主库**完全匹配不上**，只能退化到名称关键词兜底（QU-16 因此被判成 IO）。
    """
    s = str(v or "").strip()
    if not s:
        return ""
    for sep in ("/", "／", "\\", "|"):
        if sep in s:
            s = s.split(sep)[0].strip()
            break
    # 去空白：清单写「Allen & Heath」，主库存「ALLEN&HEATH」
    return "".join(s.split())


def is_section_heading(row: dict) -> bool:
    """分组标题行判定：品牌、型号、数量三列全空。

    清单常用「扬声器系统」「处理及周边设备」「3F会议室」这类小标题分节。
    它们名称常含音频关键词（"扬声器"/"处理"/"音源"），会被名称兜底成
    设备；但只要**没有数量**就不是采购项，据此与真实设备区分。
    """
    def _empty(v):
        return v is None or str(v).strip() == ""
    return (_empty(row.get("品牌")) and _empty(row.get("型号"))
            and _empty(row.get("数量")))


def _normalize_header(header) -> dict:
    """返回 {原始列下标: 标准列名}；无法识别的列忽略。"""
    m = {}
    for i, h in enumerate(header):
        key = str(h or "").strip().lower()
        if not key:
            continue
        for std, aliases in HEADER_ALIASES.items():
            if key in aliases:
                m[i] = std
                break
    return m


def _sheet_rows(ws):
    """单个工作表 -> 归一化行 dict 列表；非配置清单（无表头/价目表）返回 None。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    hi = _find_header_row(rows)
    if hi < 0:
        return None
    hmap = _normalize_header(rows[hi])
    # ★ hmap 是 {列下标: 标准名}，判存在要查 values() 而不是 key
    stds = set(hmap.values())
    if "设备名称" not in stds or "型号" not in stds:
        return None
    cn = next(i for i, k in hmap.items() if k == "设备名称")
    cq = next((i for i, k in hmap.items() if k == "数量"), None)
    # 价目表识别：列数爆炸，或数量列几乎全空/零
    if ws.max_column > PRICELIST_MAX_COLS:
        return None
    named = [r for r in rows[hi + 1:]
             if cn < len(r) and r[cn] is not None and str(r[cn]).strip()]
    if len(named) >= PRICELIST_MIN_ROWS and cq is not None:
        zero = 0
        for r in named:
            q = r[cq] if cq < len(r) else None
            if q is None or str(q).strip() == "":
                zero += 1
            else:
                try:
                    if float(q) == 0:
                        zero += 1
                except (ValueError, TypeError):
                    pass
        if zero / len(named) > PRICELIST_ZERO_RATIO:
            return None
    out = []
    for r in rows[hi + 1:]:
        if all(c is None for c in r):
            continue
        row = {}
        for i, std in hmap.items():
            if i < len(r):
                row[std] = r[i]
        for std in HEADER_ALIASES:
            row.setdefault(std, None)
        if not str(row.get("设备名称") or "").strip():
            continue
        out.append(row)
    return out


def read_xlsx_sheets(path: str) -> dict:
    """返回 {工作表名: 归一化行列表}，已排除价目表/无表头 sheet。"""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    return {ws.title: r for ws in wb.worksheets
            if (r := _sheet_rows(ws))}


def read_xlsx_rows(path: str, sheet=None) -> list:
    """读 xlsx 的设备行。

    sheet 为 None 时合并**全部有效配置 sheet**（多房间 / 多方案清单常见），
    每行打 ``_sheet`` 标记来源；指定 sheet 名时只读该表。
    """
    sheets = read_xlsx_sheets(path)
    if sheet is not None:
        return [dict(r) for r in sheets.get(sheet, [])]
    out = []
    for name, rows in sheets.items():
        for r in rows:
            d = dict(r)
            d["_sheet"] = name
            out.append(d)
    return out


# ★ 2026-09-01：这些品牌的「无线话筒套装」整体出图、**不拆分**内部话筒。
# 阳哥确认 IPS 按「套」只画一台接收机本体（含 MIX/XLR 出与 RF 入），
# 不再把内部话筒展开成独立 WIRELESS_MIC 设备。其余品牌（AUDIX / SENNHEISER 等）
# 仍照常拆分。新增「不拆分」品牌时在此加一行即可，并补一条守卫测试。
SET_EXPAND_SKIP_BRANDS = {"IPS"}


def _expand_sets(av: list) -> list:
    """把「套装」条目拆成图上应有的独立设备。

    触发条件：主库 params 含 ``set_expand``，形如 ``{"rx": 1, "tx": 2}``
    （1 台接收机 + 2 支发射端）。数量按套数成倍展开：
    清单 4 套 AUDIX AP62 → 4 台 WIRELESS_RX + 8 支 WIRELESS_MIC。

    发射端是无线设备，图上只画本体与空中 RF 口，不产生线缆连接。

    ``SET_EXPAND_SKIP_BRANDS`` 中的品牌整体出图，不做拆分。
    """
    out = []
    for e in av:
        if str(e.get("brand", "")).upper() in SET_EXPAND_SKIP_BRANDS:
            # IPS 等品牌：套装整体作为一台接收机出图，不展开内部话筒
            out.append(e)
            continue
        se = (e.get("params") or {}).get("set_expand")
        if not isinstance(se, dict):
            out.append(e)
            continue
        qty = int(e.get("quantity", 1) or 1)
        n_rx = int(se.get("rx", 0) or 0)
        n_tx = int(se.get("tx", 0) or 0)
        if not (n_rx or n_tx):
            out.append(e)
            continue
        params = {k: v for k, v in (e.get("params") or {}).items()
                  if k != "set_expand"}
        common = {k: v for k, v in e.items()
                  if k not in ("category", "quantity", "params", "features", "name")}
        if n_rx:
            r = dict(common)
            r["category"] = "WIRELESS_RX"
            r["name"] = e.get("name") or ""
            r["quantity"] = qty * n_rx
            r["params"] = dict(params)
            r["features"] = list(e.get("features") or [])
            out.append(r)
        if n_tx:
            t = dict(common)
            t["category"] = "WIRELESS_MIC"
            t["name"] = (e.get("name") or "") + "·话筒"
            t["quantity"] = qty * n_tx
            t["params"] = {}
            t["features"] = []
            out.append(t)
    return out


def apply_category_fallback(entries: list):
    """给缺类别的条目补类别；主库后置且名称也认不出的直接排除。

    返回 ``(kept, dropped_rows)``。**就地改**传入的条目（补 ``category``）。

    ★ 2026-09-01 R12：这个函数此前是 ``build_entries`` 里的一段内联代码，
      只有 **xlsx 路径**会跑到。而 Web 端第⑤步是把 `/api/parse` 归一化的 CSV
      文本回传给 ``/api/run`` → ``parse_bom`` → **原始行**，压根不走这里。
      实测后果：文本 BOM 里「设备类型」列留空的未知型号，出图时
      ``category = ""`` → 0 个端口 + ``ERROR:UNKNOWN_TYPE``，
      而不是 xlsx 路径下的 IO 兜底。**同一个清单两种下场**，取决于入口。

    顺序很重要：必须在 ``resolve_products`` **之后**调用。主库/内置库命中时
    ``enrich`` 只在 ``category`` 为空时才填（``if not entry.get("category")``），
    先跑本函数会把未知型号一律填成 IO，之后 enrich 就再也不覆盖它了。
    """
    kept, dropped = [], []
    for e in entries:
        cat = e.get("category")
        if not cat:
            cat = classify_category(e.get("name", ""))
            if not cat:
                # 主库明确后置（配件/线缆/非音频，如充电箱、中继器、主缆）
                # 且名称也命中不到任何音频类别 -> 与吊架同等处理：排除出图。
                # ★ 注意不能一刀切：QU-16 这类「主库延迟」型号要靠名称兜底成 MIXER，
                #   只有名称也识别不了时才排除；未命中主库的条目仍保留 IO 兜底。
                if str(e.get("_resolved", "")).startswith("eko-deferred"):
                    dropped.append({"设备名称": e.get("name") or e.get("model") or ""})
                    continue
                cat = "IO"
            e["category"] = cat
        kept.append(e)
    return kept, dropped


def build_entries(path: str, sheet=None):
    """读 xlsx -> 规范化条目（已排除吊架/占位项/不出图型号）。返回 (entries, dropped_rows)。

    sheet 为 None 时合并全部有效配置 sheet（多房间 / 多方案清单）；指定时只读该表。

    entries 元素含：brand/model/name/quantity/category/features(list)/params(dict)/
    electrical(dict,仅功放)/spec(原始指标参数，仅内部用)/_sheet(来源工作表名)。
    """
    raw = read_xlsx_rows(path, sheet=sheet)
    av, dropped = [], []
    for r in raw:
        name = r.get("设备名称") or r.get("名称") or r.get("name") or ""
        if is_rigging(name):
            dropped.append(r)
            continue
        # 占位项（无品牌无型号的「自配 / 按需保留」等）：直接排除，不进第②步
        if is_placeholder(name, r):
            dropped.append(r)
            continue
        # 分组标题行：清单常用「扬声器系统」「处理及周边设备」「3F会议室」
        # 这类小标题分节，品牌/型号/数量三列全空。它们不含数量，据此与
        # 真实设备区分（真实设备哪怕写「1 批」也有数量）。
        if is_section_heading(r):
            continue
        # 造价/计价清单里大量「项目特征描述」续行、小计行、空行：设备名称为空
        # 一律跳过，否则会被当成无名设备导入（实测一份清单多出 150+ 条空条目）。
        if not str(name).strip():
            continue
        av.append({
            # 品牌去国别后缀（「IPS/深圳」->「IPS」），否则主库匹配失效
            "brand": _clean_brand(r.get("品牌") or r.get("brand")),
            "model": r.get("型号") or r.get("model") or "",
            "name": name,
            "quantity": _qty(r.get("数量") or r.get("qty")),
            # 计量单位：用于「成对销售」设备的数量展开（如 UM2000AP 单位=对 → 2 支）
            "_unit": str(r.get("单位") or r.get("unit") or "").strip(),
            "spec": r.get("指标参数") or r.get("参数") or r.get("spec") or "",
            "_sheet": r.get("_sheet") or "",
        })
    # 主库补全（品牌+型号）
    resolve_products(av)
    # 成对销售设备展开：单位写成「对/副/pair」时按 2 支计
    # （IPS UM2000AP / UM2000AT 主库 remark：「1 对 = 2 支」）
    for e in av:
        if str(e.get("_unit", "")).strip().lower() in PAIR_UNITS:
            e["quantity"] = int(e.get("quantity", 1) or 1) * PAIR_UNIT_FACTOR
    # 套装拆分：清单按「套」计价，图上要分别画出接收机与发射端
    # （AUDIX / SENNHEISER 等）。★ IPS 套装整体出图、不拆分（见 _expand_sets / SET_EXPAND_SKIP_BRANDS）
    av = _expand_sets(av)
    # 主库显式标记 drawable=false（停产型号 / 视频网传 / 线缆等）：直接排除，
    # 不再走下面的名称兜底——否则「天线延长线」会被兜底成 ANTENNA。
    for e in av:
        if e.get("_no_draw"):
            dropped.append({"设备名称": e.get("name") or e.get("model") or ""})
    av = [e for e in av if not e.get("_no_draw")]
    av, dropped2 = apply_category_fallback(av)
    dropped.extend(dropped2)
    for e in av:
        cat = e.get("category")
        spec = e.get("spec", "")
        feats = set(str(x).lower() for x in e.get("features", []) or [])
        feats |= extract_features(spec, e.get("name", ""), cat)
        e["features"] = sorted(feats)
        extracted = extract_params(spec, cat)
        params = dict(extracted)
        params.update(e.get("params", {}) or {})  # 主库值优先
        e["params"] = params
        if cat == "AMP":
            elec = dict(e.get("electrical", {}) or {})
            pw = extracted.get("power_w_per_ch")
            if pw and "power_w_per_ch" not in elec:
                elec["power_w_per_ch"] = pw
            elec.setdefault("min_load_ohm", 4)
            e["electrical"] = elec
        if "有源" in (e.get("name", "") or ""):
            e["active"] = True
    return av, dropped


def to_bom_csv(entries: list) -> str:
    """规范化条目 -> 引擎 BOM CSV 文本（含类别/特性/参数/电气）。"""
    cols = ["设备类型", "品牌", "型号", "名称", "数量", "特性", "参数", "冗余",
            "处理器功能", "有源", "电气"]
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=cols)
    w.writeheader()
    for e in entries:
        params = e.get("params", {}) or {}
        elec = e.get("electrical", {}) or {}
        w.writerow({
            "设备类型": e.get("category", ""),
            "品牌": e.get("brand", ""),
            "型号": e.get("model", ""),
            "名称": e.get("name", ""),
            "数量": e.get("quantity", 1),
            "特性": ";".join(e.get("features", []) or []),
            # 复杂参数（ports_override / slots 等 list/dict）整体转 JSON，
            # 否则 CSV 往返会退化成 Python repr 字符串，后端遍历时拿到字符而非字典。
            "参数": dump_params(params),
            "冗余": e.get("redundancy", "") or "",
            "处理器功能": e.get("proc_func", "") or "",
            "有源": "是" if e.get("active") else "",
            "电气": json.dumps(elec, ensure_ascii=False) if elec else "",
        })
    return buf.getvalue()
