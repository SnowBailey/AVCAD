"""清单解析：Excel/CSV -> 规范化条目列表。支持中英文列名别名。"""
from __future__ import annotations
import csv
import json
import os

COLUMN_ALIAS = {
    "category": ["category", "设备类型", "类型", "type"],
    "brand": ["brand", "品牌"],
    "model": ["model", "型号"],
    "name": ["name", "名称", "设备名称"],
    "quantity": ["quantity", "数量", "qty"],
    "features": ["features", "特性", "功能"],
    "params": ["params", "参数", "规格"],
    "redundancy": ["redundancy", "冗余", "主备"],
    "pair": ["pair", "配对", "pair"],
    "active": ["active", "有源", "有源扬声器"],
    "proc_func": ["proc_func", "处理器功能", "dsp位置"],
    "electrical": ["electrical", "电气", "电气参数"],
    "uid": ["uid", "编号", "id"],
}


def _norm_header(h: str) -> str:
    return str(h).strip().lower()


def _map_row(row: dict) -> dict:
    out = {}
    low = {_norm_header(k): v for k, v in row.items()}
    for key, aliases in COLUMN_ALIAS.items():
        for a in aliases:
            if a.lower() in low and low[a.lower()] not in (None, ""):
                out[key] = low[a.lower()]
                break
    # 有源标志：中文/英文布尔归一化
    if "active" in out:
        out["active"] = _coerce(out["active"])
    # 参数解析：支持 JSON 或 key=value;key2=value2
    if "params" in out and out["params"]:
        out["params"] = _parse_params(out["params"])
    else:
        out["params"] = {}
    # 电气参数解析：JSON 或 key=value（功放功率/最低负载等，随清单流转）
    if "electrical" in out and out["electrical"]:
        out["electrical"] = _parse_params(out["electrical"])
    else:
        out.pop("electrical", None)
    # 数量
    try:
        out["quantity"] = int(float(out.get("quantity", 1) or 1))
    except (ValueError, TypeError):
        out["quantity"] = 1
    return out


def _parse_value(v):
    """单个参数值解析：JSON 数组/对象还原为 list/dict，其余按标量。

    ★ 关键：list/dict 必须原样保留，不能被 _coerce 压成字符串。
    否则 ports_override / slots 这类结构化参数会退化成 str，
    expand_instance 遍历时得到的是字符而不是字典，直接
    'str' object has no attribute 'get'。
    """
    s = (v if isinstance(v, str) else str(v)).strip()
    if s[:1] in ("[", "{"):
        try:
            return json.loads(s)
        except Exception:
            pass
    return _coerce(s)


def _parse_params(raw):
    if isinstance(raw, dict):
        return {k: (v if isinstance(v, (list, dict)) else _coerce(v))
                for k, v in raw.items()}
    s = str(raw).strip()
    if s.startswith("{"):
        try:
            loaded = json.loads(s)
            if isinstance(loaded, dict):
                return {k: (v if isinstance(v, (list, dict)) else _coerce(v))
                        for k, v in loaded.items()}
        except Exception:
            pass
    out = {}
    for part in s.split(";"):
        part = part.strip()
        if not part or "=" not in part:
            continue
        k, v = part.split("=", 1)
        out[k.strip()] = _parse_value(v.strip())
    return out


def dump_params(params: dict) -> str:
    """参数列序列化：含 list/dict 的复杂参数整体转 JSON（保证往返不丢类型）。

    纯标量仍用 k=v;k2=v2，保持 CSV 可读性与历史兼容。
    """
    if not params:
        return ""
    if any(isinstance(v, (list, dict)) for v in params.values()):
        return json.dumps(params, ensure_ascii=False)
    return ";".join(f"{k}={v}" for k, v in params.items())


def _coerce(v):
    if isinstance(v, bool):
        return v
    s = str(v).strip()
    if s.lower() in ("true", "yes", "是", "有"):
        return True
    if s.lower() in ("false", "no", "否", "无"):
        return False
    try:
        return int(float(s))
    except ValueError:
        try:
            return float(s)
        except ValueError:
            return s


def read_csv(path: str) -> list:
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [_map_row(r) for r in reader]


def read_xlsx(path: str) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        out.append(_map_row({header[i]: r[i] for i in range(len(header))}))
    return out


def parse_bom(path: str) -> list:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return read_xlsx(path)
    return read_csv(path)
