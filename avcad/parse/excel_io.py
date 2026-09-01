"""读取 .xls / .xlsx / .xlsm 工作簿为归一化原始行。

- ``.xlsx`` / ``.xlsm``：openpyxl（已是 AVCAD 依赖，``data_only=True`` 取公式缓存值）。
- ``.xls``：Excel 97-2003 二进制（OLE2）格式，openpyxl 不支持，必须走 xlrd。

返回结构统一为 ``{工作表名: [行, ...]}``，每行是「单元格原始值」的列表，
与 openpyxl 的 ``iter_rows(values_only=True)`` 输出一致，便于上层复用同一套
表头识别 / 价目表过滤逻辑。
"""
from __future__ import annotations

import os


def _rows_from_openpyxl(path: str) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    return {ws.title: [list(r) for r in ws.iter_rows(values_only=True)]
            for ws in wb.worksheets}


def _rows_from_xlrd(path: str) -> dict:
    import xlrd
    book = xlrd.open_workbook(path)
    out = {}
    for sh in book.sheets():
        rows = []
        for r in range(sh.nrows):
            row = []
            for c in range(sh.ncols):
                v = sh.cell_value(r, c)
                # xlrd 把空单元格给 ''；统一成 None 以对齐 openpyxl 行为
                row.append(v if v != "" else None)
            rows.append(row)
        out[sh.name] = rows
    return out


def read_workbook_sheets(path: str) -> dict:
    """返回 ``{工作表名: 原始行(list[list])}``，行内值为单元格原始值。

    支持 ``.xls`` / ``.xlsx`` / ``.xlsm``；其他扩展名按 openpyxl 处理（交由调用方报错）。
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        return _rows_from_xlrd(path)
    return _rows_from_openpyxl(path)
