"""富集层导入测试：类别推断 / 吊架排除 / 功放参数归一化 / 真实 xlsx 端到端。"""
from __future__ import annotations
import os

from avcad.workflow.importers import (
    classify_category, is_rigging, is_placeholder,
    extract_params, extract_features,
    build_entries, to_bom_csv,
)

REAL_XLSX = "/Users/mac/Desktop/测试.xlsx"


def test_is_placeholder_needs_empty_brand_and_model():
    """占位词命中还不够——必须 brand 与 model 同时为空才排除。

    否则「自配机柜」这类带型号的真实设备会被误杀。
    """
    assert is_placeholder("按需保留", {"品牌": "", "型号": ""})
    assert is_placeholder("自配", {"品牌": "", "型号": ""})
    # 有品牌或型号 -> 不排除
    assert not is_placeholder("自配", {"品牌": "IPS", "型号": ""})
    assert not is_placeholder("自配", {"品牌": "", "型号": "XX-1"})
    # 非占位词
    assert not is_placeholder("数字调音台", {"品牌": "", "型号": ""})


def test_classify_category_basic():
    assert classify_category("数字功率放大器") == "AMP"
    assert classify_category("数字调音台") == "MIXER"
    assert classify_category("数字音频处理器（带网络监测及dante）") == "PROCESSOR"
    assert classify_category("16mm大振膜方杆模拟话筒") == "SOURCE"
    assert classify_category("左右全频线阵列扬声器") == "SPEAKER"


def test_is_rigging_excluded():
    assert is_rigging("线阵列扬声器吊架")
    assert is_rigging("ML210 飞行架")
    assert not is_rigging("左右全频线阵列扬声器")
    assert classify_category("线阵列扬声器吊架") is None


def test_extract_params_amp():
    spec = "1、8Ω 2通道≥：900W \n2、信噪比≥108dB"
    p = extract_params(spec, "AMP")
    assert p["channels"] == 2
    assert p["power_w_per_ch"] == 900


def test_extract_features_dante_control():
    feats = extract_features("具备Dante冗余网口；支持集控系统控制", "数字音频处理器", "PROCESSOR")
    assert "dante" in feats
    assert "control" in feats
    # analog 仅在规格含「模拟」时抽取
    feats2 = extract_features("具备Dante；模拟输入12路/模拟输出8路", "数字音频处理器", "PROCESSOR")
    assert "analog" in feats2


def test_to_bom_csv_roundtrip_electrical():
    entries = [{
        "category": "AMP", "brand": "ezacoustics", "model": "EM30D",
        "name": "数字功率放大器", "quantity": 5,
        "features": ["analog", "control"], "params": {"channels": 2},
        "electrical": {"min_load_ohm": 4, "power_w_per_ch": 900},
    }]
    csv_text = to_bom_csv(entries)
    assert "电气" in csv_text
    # 解析回来电气应保留
    from avcad.parse.bom_parser import parse_bom
    import tempfile
    with tempfile.NamedTemporaryFile("w", suffix=".csv", delete=False, encoding="utf-8") as f:
        f.write(csv_text); tmp = f.name
    try:
        parsed = parse_bom(tmp)
        assert parsed[0]["electrical"]["power_w_per_ch"] == 900
        assert parsed[0]["electrical"]["min_load_ohm"] == 4
        assert parsed[0]["params"]["channels"] == 2
    finally:
        os.unlink(tmp)


def test_to_bom_csv_roundtrip_ports_override():
    """★ 回归：ports_override（list of dict）经 CSV 往返必须仍是 list。

    历史 bug：to_bom_csv 用 str() 把 list 写成 Python repr，回读后变成字符串，
    expand_instance 遍历它得到**单个字符**，抛
    'str' object has no attribute 'get'（IPS CF6300 / CF6300WB 触发）。
    """
    override = [
        {"name": "PHX", "side": "right", "signal": "XLR", "role": "out",
         "label": "PHX", "count": 4},
        {"name": "MIX", "side": "right", "signal": "XLR", "role": "out",
         "label": "MIX", "count": 1},
    ]
    entries = [{
        "category": "IO", "brand": "IPS", "model": "CF6300",
        "name": "有线无线融合会议主机", "quantity": 1,
        "features": [], "params": {"ports_override": override},
    }]
    csv_text = to_bom_csv(entries)
    from avcad.parse.bom_parser import parse_bom
    import tempfile
    with tempfile.NamedTemporaryFile("w", suffix=".csv", delete=False, encoding="utf-8") as f:
        f.write(csv_text); tmp = f.name
    try:
        parsed = parse_bom(tmp)
        got = parsed[0]["params"]["ports_override"]
        assert isinstance(got, list), f"ports_override 退化为 {type(got).__name__}"
        assert len(got) == 2 and got[0]["count"] == 4
    finally:
        os.unlink(tmp)


def test_scalar_params_still_readable():
    """纯标量参数仍用 k=v;k2=v2，保持 CSV 可读性（不整体转 JSON）。"""
    entries = [{
        "category": "MIXER", "brand": "Yamaha", "model": "TF5",
        "name": "调音台", "quantity": 1, "features": [],
        "params": {"inputs": 32, "outputs": 16},
    }]
    csv_text = to_bom_csv(entries)
    assert "inputs=32;outputs=16" in csv_text


def test_build_entries_skips_empty_name_rows(tmp_path):
    """★ 回归：造价清单的「项目特征描述」续行/空行不应被当成设备导入。

    真实清单 智慧剧场20260728V2.xlsx 有 181 行，其中 150+ 行设备名称为空，
    未过滤时会产生大量无名 IO 条目。
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    # 真实造价清单结构：设备名称/品牌/型号/数量 与 项目特征描述 分处不同列区
    ws.append(["项目名称", "项目特征描述", None, None, "设备名称", "品牌", "型号", "数量"])
    ws.append(["数字调音台", "1.名称：数字调音台", None, None, "数字调音台", "IPS", "TEST-M1", 2])
    # 特征描述续行：设备名称列为空
    ws.append([None, "；LED封装形式：SMD1515黑灯；点间距≤1.86mm", None, None, None, None, None, None])
    ws.append([None, None, None, None, None, None, None, None])   # 空行
    ws.append([None, None, None, None, "", "", "", ""])            # 空串行
    path = tmp_path / "empty_rows.xlsx"
    wb.save(path)
    entries, dropped = build_entries(str(path))
    assert len(entries) == 1, f"应只剩 1 条，实际 {len(entries)}: {[e.get('name') for e in entries]}"
    assert entries[0]["model"] == "TEST-M1"


def test_deferred_keeps_name_fallback_but_drops_accessories(tmp_path):
    """★ 回归：主库后置条目的两种去向必须分开。

    - QU-16：主库 deferred，但名称「数字调音台」命中关键词 -> 兜底成 MIXER（必须保留）
    - CF6300WCB 充电箱：主库 deferred 且名称无音频关键词 -> 排除出图（与吊架同等待遇）

    历史 bug：一律兜底成 "IO"，导致 category=None 的「不出图」约定形同虚设，
    充电箱/中继器/线缆都被画成 IO 设备。
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["设备名称", "品牌", "型号", "数量"])
    ws.append(["数字调音台", "ALLEN&HEATH", "QU-16", 1])
    ws.append(["无线单元充电箱", "IPS", "CF6300WCB", 1])
    path = tmp_path / "deferred.xlsx"
    wb.save(path)
    entries, dropped = build_entries(str(path))
    models = {str(e.get("model", "")).upper(): e for e in entries}
    assert "QU-16" in models, "QU-16 应靠名称兜底保留"
    assert models["QU-16"]["category"] == "MIXER"
    assert "CF6300WCB" not in models, "充电箱不传音频，应被排除"
    dropped_names = [str(d.get("设备名称") or "") for d in dropped]
    assert any("充电箱" in n for n in dropped_names)


THEATRE_XLSX = "/Users/mac/Desktop/202601/智慧剧场20260728V2.xlsx"


def test_real_theatre_xlsx_end_to_end():
    """★ 真实造价清单端到端回归（本 bug 的原始复现文件）。

    覆盖两件事：
    1. 大量「项目特征描述」续行/空行不被当成设备导入；
    2. IPS CF6300 / CF6300WB 的 ports_override 经 CSV 往返仍是 list，
       出图不再抛 'str' object has no attribute 'get'。
    """
    import json
    from avcad.ui.app import _dispatch
    if not os.path.exists(THEATRE_XLSX):
        import pytest
        pytest.skip("智慧剧场20260728V2.xlsx 不在本机")
    entries, _ = build_entries(THEATRE_XLSX)
    assert len(entries) < 60, f"空行未过滤，条目数={len(entries)}"
    csv_text = to_bom_csv(entries)
    r = _dispatch("/api/run", json.dumps({"bom": csv_text, "name": "智慧剧场"}))
    assert "error" not in r, f"出图报错: {r.get('error')}"
    assert r["validation"]["ok"] is True
    # CF6300(2 项 override 展开为 5 口) / CF6300WB(6 项) 端口正常
    by_model = {}
    for d in r["devices"]:
        by_model.setdefault(d["model"], len(d["ports"]))
    assert by_model.get("CF6300", 0) > 0
    assert by_model.get("CF6300WB", 0) > 0


def test_build_entries_real_xlsx():
    if not os.path.exists(REAL_XLSX):
        import pytest
        pytest.skip("真实清单 测试.xlsx 不在本机")
    entries, dropped = build_entries(REAL_XLSX)
    # 吊架被排除
    dropped_names = [str(d.get("设备名称") or d.get("名称")) for d in dropped]
    assert any("吊架" in n for n in dropped_names)
    by_model = {e["model"]: e for e in entries}
    # QU-16 主库 deferred -> 兜底 MIXER
    assert by_model["QU-16"]["category"] == "MIXER"
    # 功放归一化
    assert by_model["EM30D"]["params"].get("channels") == 2
    assert by_model["EM30D"]["electrical"]["power_w_per_ch"] == 900
    assert by_model["EM50Q"]["params"].get("channels") == 4
    # 扬声器数量合计
    spk = sum(e["quantity"] for e in entries if e["category"] == "SPEAKER")
    assert spk == 28
    # ML210FB 不在条目内
    assert "ML210FB" not in by_model


def _xlsx(tmp_path, rows):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["设备名称", "品牌", "型号", "数量"])
    for r in rows:
        ws.append(list(r))
    path = tmp_path / "ips_catalog.xlsx"
    wb.save(path)
    return str(path)


def test_ips_catalog_round3_nodraw_and_reclass(tmp_path):
    """主库矫正第三轮：非设备/配件不出图，误分类设备归位。"""
    path = _xlsx(tmp_path, [
        # —— 应排除（no_draw）——
        ("公－母5米屏蔽主缆", "IPS", "CF-C1132-05", 4),   # 线缆
        ("无线单元充电箱", "IPS", "CF6300WCB", 1),        # 充电箱
        ("Aries 控制面板（白色）", "IPS", "CP-1PW", 2),    # 控制面板
        ("线阵列扬声器飞行架", "IPS", "ML210FB", 2),       # 安装架
        ("鹅颈麦克风咪杆", "IPS", "CF2223", 10),         # 咪杆（配件）
        # —— 应保留且分类纠正 ——
        ("智能自动混音器", "IPS", "AM860", 1),            # SOURCE -> MIXER
        ("单电脑DI盒", "IPS", "SI Box", 1),              # SOURCE -> IO
        ("双电脑DI盒", "IPS", "DI Box", 1),              # SOURCE -> IO
        # —— 占位项（无品牌无型号）——
        ("按需保留", "", "", 1),
        ("自配", "", "", 1),
    ])
    entries, dropped = build_entries(path)
    got = {str(e.get("model")): e for e in entries}
    dropped_names = [str(d.get("设备名称") or d.get("名称")) for d in dropped]

    # 排除
    for m in ("CF-C1132-05", "CF6300WCB", "CP-1PW", "ML210FB", "CF2223"):
        assert m not in got, f"{m} 应被 no_draw 排除"
    # 占位项也排除，且不用用户手动点「不需要」
    assert any("按需保留" in n for n in dropped_names)
    assert any("自配" in n for n in dropped_names)

    # 分类纠正
    # AM860 是**自动混音器**（阳哥 2026-08-31 订正：8 进 1 出，可级联），
    # 不是音频处理器——此前按命名推断成了 PROCESSOR。
    assert got["AM860"]["category"] == "MIXER"
    assert (got["AM860"].get("params") or {}).get("cascade"), "AM860 应可级联"
    # DI 盒 = 电脑非平衡音频转平衡进调音台，属**音源侧**；
    # 归 IO 会被画成「调音台 -> DI盒」（chain 里 IO 在 MIXER 之后），方向反了。
    assert got["SI Box"]["category"] == "SOURCE"
    assert got["DI Box"]["category"] == "SOURCE"


def test_ezacoustics_accessories_nodraw(tmp_path):
    """ezacoustics 配件不再混进设备类。

    特别关注名称含「扬声器 / 音箱」的配件——它们会被 CATEGORY_KW 兜底成
    SPEAKER，必须靠主库 no_draw 压住，光靠关键词排除不干净。
    """
    path = _xlsx(tmp_path, [
        ("监听扬声器配件", "ezacoustics", "W12MU", 4),      # 名称含"扬声器"
        ("音箱支架连接适配件", "ezacoustics", "AQUA-ADAP", 2),  # 名称含"音箱"
        ("线阵列扬声器吊挂架", "ezacoustics", "T12FB", 2),
        ("25米4芯音箱连接线", "ezacoustics", "TS音箱连接线", 6),
        ("线箱", "ezacoustics", "TS线箱", 1),               # 现归 AMP
        ("Console-Link DM7 镜像控制器", "ezacoustics", "Console-Link DM7", 1),
        # —— 应保留 ——
        ("12\"两分频同轴舞台监听扬声器", "ezacoustics", "W12M", 4),
        ("沉浸声音频渲染服务器", "ezacoustics", "X-CORE S", 1),
        ("数字音频接入盒", "ezacoustics", "RDD12", 1),
    ])
    entries, _ = build_entries(path)
    got = {str(e.get("model")): e for e in entries}

    for m in ("W12MU", "AQUA-ADAP", "T12FB", "TS音箱连接线",
              "TS线箱", "Console-Link DM7"):
        assert m not in got, f"{m} 应被 no_draw 排除"

    assert got["W12M"]["category"] == "SPEAKER"
    assert got["X-CORE S"]["category"] == "PROCESSOR"
    assert got["RDD12"]["category"] == "IO"


def test_ips_discontinued_units_nodraw_and_cf6300wb_antenna(tmp_path):
    """阳哥裁定：停产主机的配套单元一并停产；CF6300WB 归 ANTENNA。

    CF63 系列主机 CF6300 在产，其单元必须保留；CF68/CF68W/CF62/CF61 的
    主机均已停产，配套单元同停。
    """
    path = _xlsx(tmp_path, [
        ("数字会议单元", "IPS", "CF6821", 4),        # CF68  主机停产
        ("无线数字会议单元", "IPS", "CF6850", 4),     # CF68W 主机停产
        ("数字会议单元", "IPS", "CF6211", 4),        # CF62  主机停产
        ("数字会议单元", "IPS", "CF6110L", 2),       # CF61  主机停产
        ("数字会议单元", "IPS", "CF6310", 4),        # CF63  在产 -> 保留
        ("鹅颈数字会议单元", "IPS", "CF6319L", 4),    # CF63  在产 -> 保留
        ("无线会讨天线盒", "IPS", "CF6300WB", 1),     # -> ANTENNA
        ("有线无线融合会议主机", "IPS", "CF6300", 1),  # -> MIC_HOST
    ])
    entries, _ = build_entries(path)
    got = {str(e.get("model")): e for e in entries}

    for m in ("CF6821", "CF6850", "CF6211", "CF6110L"):
        assert m not in got, f"{m} 主机已停产，单元应一并排除"

    assert got["CF6310"]["category"] == "SOURCE"
    assert got["CF6319L"]["category"] == "SOURCE"
    assert got["CF6300"]["category"] == "MIC_HOST"
    assert got["CF6300WB"]["category"] == "ANTENNA"
