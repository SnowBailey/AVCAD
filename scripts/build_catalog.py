#!/usr/bin/env python3
"""易科国际产品资料清单 -> AVCAD 型号库（音频 V1）。

用法:
  python scripts/build_catalog.py <清单.xlsx> [--out avcad/avcad/data/eko_catalog.json]

- 按品牌分表解析；表头行 = 含「设备名称」的行
- 产品行 = 「物料编码/产品编码」非空
- 以 序号 列中的 组/分段 标题做类别上下文
- 仅识别音频类别（V1）；中控(V2)/视频(V3)/电源/灯光 标记后置
输出: eko_catalog.json（所有产品 + 分类 + 特性 + 参数 + 可信度）
"""
from __future__ import annotations
import argparse, json, os, re, sys

import openpyxl

AUDIO_CATS = ["SOURCE", "WIRELESS_MIC", "ANTENNA", "ANT_DIST", "ANT_COMBINE",
              "WIRELESS_RX",
              "MIXER", "PROCESSOR", "SPEAKER_MGR", "AMP", "SPEAKER", "SWITCH", "IO"]

HEADER_ALIASES = {
    "seq": ["序号", "编号", "NO", "No", "no", "项号"],
    "name": ["设备名称", "产品名称", "名称"],
    "brand": ["品牌"],
    "model": ["型号", "厂家型号"],
    "desc": ["产品描述", "规格", "规格参数", "描述"],
    "country": ["国别", "国家", "产地", "属地"],
    "qty": ["数量"],
    "unit": ["单位"],
    "code": ["物料编码", "产品编码", "编码", "货号", "物料代码", "ERP内部型号"],
    "price": ["单价"],
    "total": ["合价"],
    "remark": ["备注", "说明", "注意事项", "备货情况"],
}

KNOWN_WIRELESS = {"SHURE", "SENNHEISER", "AUDIO-TECHNICA", "LECTROSONICS", "SONY"}

# 仅当关键词未命中且名称不含配件词时，用品牌做兜底归类
BRAND_CAT = {
    "EAW": "SPEAKER", "COMMUNITY": "SPEAKER", "L-ACOUSTICS": "SPEAKER",
    "PAN ACOUSTICS": "SPEAKER", "APART": "SPEAKER", "KARNO": "SPEAKER",
    "TECTONIC": "SPEAKER", "MACKIE": "SPEAKER", "JBL": "SPEAKER",
    "POWERSOFT": "AMP",
    "YAMAHA": "MIXER", "ALLEN&HEATH": "MIXER", "ALLEN & HEATH": "MIXER",
    "MIDAS": "MIXER", "SOUNDCRAFT": "MIXER", "DIGICO": "MIXER", "STUDER": "MIXER",
    "BEHRINGER": "MIXER", "PRESONUS": "MIXER", "AVID": "MIXER",
    "SYMETRIX": "PROCESSOR", "BSS": "PROCESSOR", "BIAMP": "PROCESSOR",
    "XILICA": "PROCESSOR", "ASHLY": "PROCESSOR", "MEDIA MATRIX": "PROCESSOR",
    "AUDINATE": "SWITCH", "NETGEAR": "SWITCH", "TELEVIC": "SOURCE",
}
# 仅用于「品牌兜底」排除的广义配件词（含功能性词，不能用于主匹配排除）
ACCESSORY_TOKENS = ["卡", "软件", "备件", "配件", "线", "缆", "吊架", "机柜", "罩",
                    "包", "法兰", "面板", "接插件", "连接器", "麦架", "话筒架",
                    "模块", "电池", "适配器", "转接头", "支架", "箱", "袋",
                    "变压器", "网关", "板", "换能", "网络", "界面", "分配器",
                    "管理器", "监测", "分析", "控制", "供电", "电源", "接口"]

# 仅在「产品名称」中出现的、明确不是可绘制核心设备的配件词。
# 命中即后置人工（避免描述里提到某设备类型而把配件误判为核心设备）。
# 注意：不含「卡/模块/箱」等合法音频扩展件（它们是真实 IO/PROCESSOR）。
ACCESSORY_NAME_TOKENS = ["配件", "备件", "软件", "吊架", "机柜", "电池", "转接头",
                         "适配器", "法兰", "接插件", "连接器", "麦架", "话筒架",
                         "袋", "变压器", "网关", "升级包", "备件包", "扩展包",
                         "软件包", "功能包", "箱包", "包装"]


NAME_HEADER = {"设备名称", "产品名称", "名称"}


def find_header(row):
    for i, v in enumerate(row):
        if v and str(v).strip() in NAME_HEADER:
            return i
    return -1


def col_index_map(header):
    """返回 {field: col_index}，按别名匹配。"""
    m = {}
    for field, aliases in HEADER_ALIASES.items():
        for ci, v in enumerate(header):
            if v and str(v).strip() in aliases:
                m[field] = ci
                break
    return m


def _kw(text, up, b):
    """纯关键词分类；text 已包含待匹配文本。返回类别或 None。"""
    if "调音台" in text:
        return "MIXER"
    if "接收机" in text and ("无线" in text or "分集" in text or "话筒" in text or b in KNOWN_WIRELESS):
        return "WIRELESS_RX"
    if "发射" in text or ("无线" in text and ("话筒" in text or "麦克风" in text)):
        return "WIRELESS_MIC"
    if "天线分配" in text or "天线分线" in text or ("分配" in text and "天线" in text):
        return "ANT_DIST"
    if "天线" in text:
        return "ANTENNA"
    if "功放" in text or "功率放大器" in text or "功率放大" in text:
        return "AMP"
    if "扬声器管理" in text or "音箱管理" in text or "音箱控制器" in text:
        return "SPEAKER_MGR"
    if "处理器" in text or "DSP" in up or "音频处理" in text or "分频器" in text \
       or "均衡器" in text or "效果器" in text or "话放" in text or "前置放大" in text \
       or "转换器" in text or "矩阵处理" in text:
        return "PROCESSOR"
    if "扬声器" in text or "音箱" in text or "音响" in text:
        return "SPEAKER"
    if "交换机" in text or "SWITCH" in up:
        return "SWITCH"
    # 音频接口 / 声卡 / interface / I/O 箱 → IO 设备。
    # 名称即语义（如「多通道声卡」= 音频接口），必须先于「话筒/会议单元」判定，
    # 否则声卡会被误判为音源(SOURCE)。
    if "声卡" in text or "音频接口" in text or "interface" in text.lower() \
       or "I/O" in up or "接口箱" in text or "IO箱" in text or "扩展接口" in text:
        return "IO"
    if "话筒" in text or "麦克风" in text or "传声器" in text or "鹅颈" in text \
       or "电容话筒" in text or "DI盒" in text \
       or "主席单元" in text or "代表单元" in text or "讨论单元" in text or "会议单元" in text:
        return "SOURCE"
    return None


# ---- 人工校正（来自阳哥的领域知识，覆盖清单未明确写出的参数/类别）----
# 键： (品牌大写, 型号大写)
MANUAL_CAT = {
    # RY16-AE 是插到 RPio 卡槽里的 AES 输入/输出卡 -> 属 IO 类（而非处理器）
    ("YAMAHA", "RY16-AE"): "IO",
    # IPS 手拉手会议主机/天线板：按音频 I/O 出图
    ("IPS", "CF6300"): "IO",
    ("IPS", "CF6300WB"): "IO",
    # IPS UM2000ASD：位于天线与 UM2000ATD(十通道分配器)之间，
    # 把两路天线 RF 合成为一路 -> 天线信号合路器（Combiner），非分配器
    ("IPS", "UM2000ASD"): "ANT_COMBINE",
    # ALLEN&HEATH 仅 QU16 出图，其余型号后置
    ("ALLEN&HEATH", "QU16"): "MIXER",
    ("ALLEN & HEATH", "QU16"): "MIXER",
    # EZACOUSTICS RDD12：双路 Dante 主备输入 + 12×XLR 输出，从 Dante 网络取信号
    ("EZACOUSTICS", "RDD12"): "IO",
}

MANUAL_PARAMS = {
    # YAMAHA CS-R10 控制台面：固定 8 模拟输入 / 8 模拟输出；
    # HY/MY 是插卡槽位，未插卡时不对外，需在模块上以卡槽形式绘制。
    ("YAMAHA", "CS-R10"): dict(
        inputs=8, outputs=8, feature_ports=False,
        slots=[{"type": "HY", "count": 4, "label": "HY"},
               {"type": "MY", "count": 2, "label": "MY"}],
        note="台面；DSP 引擎=DSP-RX；HY/MY 卡槽未插卡时不对外"),
    ("YAMAHA", "CS-R10-S"): dict(
        inputs=8, outputs=8, feature_ports=False,
        slots=[{"type": "HY", "count": 4, "label": "HY"},
               {"type": "MY", "count": 2, "label": "MY"}],
        note="台面（内置 DSP）；HY/MY 卡槽未插卡时不对外"),
    # YAMAHA RPio622 扩展 I/O 机架：完全由卡槽组成，未插卡时无对外接口
    ("YAMAHA", "RPio622"): dict(
        ports_override=[],
        slots=[{"type": "RY", "count": 6, "label": "RY"},
               {"type": "HY", "count": 2, "label": "HY"},
               {"type": "MY", "count": 2, "label": "MY"}],
        note="扩展 I/O 机架；RY/HY/MY 卡槽未插卡时不对外"),
    # YAMAHA RMio64-D：Dante/MADI 转换器
    ("YAMAHA", "RMio64-D"): dict(
        ports_override=[
            {"name": "DANTE_PRI", "side": "left", "signal": "DANTE", "role": "in", "label": "DANTE-P"},
            {"name": "DANTE_SEC", "side": "left", "signal": "DANTE", "role": "in", "label": "DANTE-S"},
            {"name": "MADI_IN_BNC", "side": "right", "signal": "AES", "role": "in", "label": "MADI-IN-BNC"},
            {"name": "MADI_OUT_BNC", "side": "right", "signal": "AES", "role": "out", "label": "MADI-OUT-BNC"},
            {"name": "MADI_IN_OPT", "side": "right", "signal": "OPTICAL", "role": "in", "label": "MADI-IN-OPT"},
            {"name": "MADI_OUT_OPT", "side": "right", "signal": "OPTICAL", "role": "out", "label": "MADI-OUT-OPT"},
            {"name": "WCLK", "side": "top", "signal": "WCLK", "role": "in", "label": "WCLK"},
        ],
        note="Dante/MADI 格式转换器；64×64 通道"),
    # EAW UX3600 扬声器处理器：3 模拟输入 / 6 模拟输出（固定 I/O，不走 zones 模板）
    ("EAW", "UX3600"): dict(
        ports_override=[
            {"name": "IN", "side": "left", "signal": "XLR", "role": "in", "label": "IN", "count": 3},
            {"name": "OUT", "side": "right", "signal": "XLR", "role": "out", "label": "OUT", "count": 6},
        ],
        note="3×XLR 输入 / 6×XLR 输出 扬声器处理器"),
    # IPS CF6804 四通道无线会议系统：1×XLR MIX out + 4×XLR 独立输出
    ("IPS", "CF6804"): dict(
        ports_override=[
            {"name": "MIX", "side": "right", "signal": "XLR", "role": "out", "label": "MIX", "count": 1},
            {"name": "OUT", "side": "right", "signal": "XLR", "role": "out", "label": "OUT", "count": 4},
        ],
        note="四通道无线会议系统；1×XLR MIX out + 4×XLR OUT"),
    # IPS CF6300 手拉手会议主机：4 路凤凰端子分区输出 + 1 路卡侬头混合输出
    ("IPS", "CF6300"): dict(
        ports_override=[
            {"name": "PHX", "side": "right", "signal": "XLR", "role": "out", "label": "PHX", "count": 4},
            {"name": "MIX", "side": "right", "signal": "XLR", "role": "out", "label": "MIX", "count": 1},
        ],
        note="手拉手会议主机；4×凤凰端子分区输出 + 1×XLR MIX 输出"),
    # IPS CF6300WB 无线会讨天线板：4×6P_DIN 音频输入 + 1×XLR 输入；通信口 NET/UPDATE/RS485/RS232
    ("IPS", "CF6300WB"): dict(
        ports_override=[
            {"name": "DIN", "side": "left", "signal": "XLR", "role": "in", "label": "DIN", "count": 4},
            {"name": "XLR", "side": "left", "signal": "XLR", "role": "in", "label": "IN", "count": 1},
            {"name": "NET", "side": "top", "signal": "IP", "role": "in", "label": "NET", "count": 1},
            {"name": "UPDATE", "side": "top", "signal": "GPIO", "role": "in", "label": "UPDATE", "count": 1},
            {"name": "RS485", "side": "top", "signal": "RS232", "role": "in", "label": "RS485", "count": 1},
            {"name": "RS232", "side": "top", "signal": "RS232", "role": "in", "label": "RS232", "count": 1},
        ],
        note="无线会讨天线板；4×6P_DIN IN + 1×XLR IN；NET/UPDATE/RS485/RS232"),
    # EZACOUSTICS RDD12 双路 Dante 接口主备：2×DANTE in(top, 主备) + 12×XLR out(right)
    ("EZACOUSTICS", "RDD12"): dict(
        ports_override=[
            {"name": "DANTE", "side": "top", "signal": "DANTE", "role": "in", "label": "DANTE", "count": 2},
            {"name": "OUT", "side": "right", "signal": "XLR", "role": "out", "label": "OUT", "count": 12},
        ],
        note="双路 Dante 主备输入 + 12×XLR 输出；从 Dante 网络取信号"),
}


def _apply_manual(p: dict) -> None:
    key = (str(p.get("brand") or "").upper(), str(p.get("model") or "").upper())
    brand = key[0]
    if key in MANUAL_CAT:
        p["category"] = MANUAL_CAT[key]
        p["defer_reason"] = None
    # ALLEN&HEATH 仅 QU16 出图，其他型号强制后置
    if brand == "ALLEN&HEATH" and key[1] != "QU16":
        p["category"] = None
        p["defer_reason"] = "ALLEN&HEATH 仅 QU16 出图"
    if key in MANUAL_PARAMS:
        params = dict(p.get("params") or {})
        params.update(MANUAL_PARAMS[key])
        p["params"] = params


def classify(name, group, section, desc, brand):
    """返回 (audio_category_or_None, reason_or_None)。

    注意：组/分段标题只作为「兜底上下文」，绝不进入主关键词匹配，
    否则会把分段标题里的词（如「功率放大器」）泄漏到产品上造成误判。
    """
    name = name or ""
    desc = desc or ""
    b = (brand or "").upper().strip()
    # 非音频（V1 后置）—— 仅看产品自身名称/描述
    full = f"{name} {desc}"
    flow = full.lower()
    fup = full.upper()
    # 注意：单纯「电源」不能判为非音频——功放规格里常见「开关电源」等。
    # 仅当品牌为 FURMAN 或明确是电源调理/时序/分配设备才后置。
    if b == "FURMAN" or "PDU" in fup or "电源调节器" in full or "电源时序" in full \
       or "时序电源" in full or "电源净化" in full or "稳压电源" in full \
       or "配电" in full or "电源分配" in full or "供电单元" in full:
        return (None, "电源/PDU(非信号)")
    if "灯具" in full or "调光" in full or "灯控" in full or "舞台灯" in full:
        return (None, "灯光(V3)")
    # 视频设备判定需严格：仅明确的视频系统词，避免把「液晶显示屏」(设备自带屏)误判为视频
    if "摄像机" in full or "投影" in full or "视频矩阵" in full or "视频分配" in full \
       or "LED显示屏" in full or "显示屏幕" in full or "拼接屏" in full or "大屏" in full or "视频墙" in full:
        return (None, "视频(V3)")
    if "内部通讯" in full or "intercom" in flow or "通话系统" in full or "无线通话" in full:
        return (None, "通讯(V2)")
    if "中控" in full or "控制主机" in full or "可编程控制" in full:
        return (None, "中控(V2)")

    # 配件/耗材/软件/安装件等明确不是可绘制核心设备 -> 直接后置人工。
    # 仅看产品名称本身（不看描述），避免描述里提到某设备类型而误判。
    if any(tok in name for tok in ACCESSORY_NAME_TOKENS):
        return (None, "音频配件(需人工)")

    # 品牌级规则：Green-GO 是无线内通系统，基本不用，不纳入音频出图
    if b == "GREEN-GO":
        return (None, "通讯(V2)")

    # 主匹配：产品名称优先（名称即设备类型，不被描述干扰）
    cat_name = _kw(name, name.upper(), b)
    if cat_name:
        return (cat_name, None)
    # 名称无明确类型时，再用 名称+描述；但「功放」仅当出现在名称中才采信，
    # 否则扬声器规格里常见的「建议功放功率」会把音箱误判为功放。
    cat = _kw(full, fup, b)
    if cat == "AMP" and "功放" not in name and "功率放大器" not in name and "功率放大" not in name:
        cat = None
    if cat:
        return (cat, None)
    # 兜底：分段/组标题上下文（仅当名称/描述无法判定时）
    ctx = f"{group or ''} {section or ''}"
    if ctx.strip():
        cat = _kw(ctx, ctx.upper(), b)
        if cat:
            return (cat, None)
    # 品牌兜底（排除明显配件词）
    if b in BRAND_CAT and not any(tok in name for tok in ACCESSORY_TOKENS):
        return (BRAND_CAT[b], None)
    return (None, "音频未识别")


def extract_features(name, desc):
    text = " ".join(str(x) for x in [name, desc] if x)
    low = text.lower()
    f = set()
    if "dante" in low:
        f.add("dante")
    if re.search(r"\baes\b", low) or "aes67" in low or "aes3" in low:
        f.add("aes")
    if "avb" in low:
        f.add("avb")
    if any(k in text for k in ["控制", "RS-232", "RS232", "TCP", "IP", "网络控制", "GPIO", "以太网"]):
        f.add("control")
    if "模拟" in text or "XLR" in text.upper():
        f.add("analog")
    if "无线" in text:
        f.add("wireless")
    if "幻象" in text or "48V" in text or "phantom" in low:
        f.add("phantom")
    return sorted(f)


def extract_params(desc):
    """尽力而为的参数提取，带 low 可信度标记。"""
    d = {}
    if not desc:
        return d
    # 阻抗（单值存标量，多值才存列表）
    imps = sorted({int(x) for x in re.findall(r"(\d+)\s*Ω", desc)})
    if len(imps) == 1:
        d["impedance_ohm"] = imps[0]
    elif len(imps) > 1:
        d["impedance_ohm"] = imps
    # 功率 W（靠近「功率」）
    m = re.search(r"功率[：: ]*\s*(\d+)\s*W", desc)
    if m:
        d["power_w"] = int(m.group(1))
    # 通道/输入/输出（取首个出现的数字，低可信）
    m = re.search(r"(\d+)\s*通道", desc)
    if m:
        d["channels_hint"] = int(m.group(1))
    m = re.search(r"(\d+)\s*路", desc)
    if m and "channels_hint" not in d:
        d["channels_hint"] = int(m.group(1))
    # BNC 天线数
    bnc = desc.upper().count("BNC")
    if bnc:
        d["bnc"] = bnc
    return d


def parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = -1
    for ri, row in enumerate(rows):
        if find_header(row) >= 0:
            hdr_idx = ri
            break
    if hdr_idx < 0:
        return []
    header = rows[hdr_idx]
    cim = col_index_map(header)
    has_code_col = "code" in cim

    def get(row, field, default=None):
        ci = cim.get(field)
        if ci is None or ci >= len(row):
            return default
        v = row[ci]
        return v

    out = []
    cur_group = None
    cur_section = None
    for row in rows[hdr_idx + 1:]:
        if not row or all(v in (None, "") for v in row):
            continue
        seq = get(row, "seq")
        s_seq = str(seq).strip() if seq is not None else ""
        name = get(row, "name") or ""
        brand = get(row, "brand") or ""
        model = get(row, "model") or ""
        code = get(row, "code")
        code_val = ""
        if code not in (None, ""):
            # 单元格可能含多个编码（换行分隔），取第一个
            code_val = str(code).split("\n")[0].split("\r")[0].strip()
        # 无编码列时以型号作为回退键
        if not has_code_col and not code_val:
            code_val = str(model).strip()

        # ---- 上下文行（无编码）----
        if not code_val:
            if re.match(r"^[A-Za-z]$", s_seq):
                cur_section = name or s_seq
            elif not re.match(r"^[A-Za-z]\d+$", s_seq):
                # 组/分段标题（如「数字调音台」，可能在 col0 或 name 列）
                cur_group = (name or s_seq).strip() or cur_group
            continue

        # ---- 产品行（编码非空）----
        desc = get(row, "desc") or ""
        cat, reason = classify(name, cur_group, cur_section, desc, brand)
        feats = extract_features(name, desc)
        params = extract_params(desc)
        active = ("有源" in name) or ("有源" in desc)
        try:
            price = float(get(row, "price") or 0)
        except (TypeError, ValueError):
            price = None
        prod = {
            "code": code_val,
            "brand": str(brand).strip(),
            "model": str(model).strip(),
            "name": str(name).strip(),
            "group": cur_group,
            "section": cur_section,
            "category": cat,
            "defer_reason": reason,
            "features": feats,
            "active": active,
            "params": params,
            "country": str(get(row, "country") or "").strip(),
            "unit": str(get(row, "unit") or "").strip(),
            "qty": get(row, "qty"),
            "price": price,
            "remark": str(get(row, "remark") or "").strip(),
            "sheet": ws.title,
        }
        _apply_manual(prod)
        out.append(prod)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # avcad/
    ap.add_argument("--out", default=os.path.join(_root, "avcad", "data", "eko_catalog.json"))
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    allp = []
    sheets_scanned = 0
    for ws in wb.worksheets:
        prods = parse_sheet(ws)
        if prods:
            sheets_scanned += 1
            allp.extend(prods)

    # 统计
    from collections import Counter
    cat_c = Counter(p["category"] or ("DEFER:" + (p["defer_reason"] or "未知")) for p in allp)
    audio = [p for p in allp if p["category"] in AUDIO_CATS]
    defer = [p for p in allp if p["category"] is None]

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump({"total": len(allp), "products": allp}, f, ensure_ascii=False, indent=1)

    print(f"扫描品牌表: {sheets_scanned} / {len(wb.sheetnames)}")
    print(f"产品总数: {len(allp)} | 音频可识别: {len(audio)} | 后置/未识别: {len(defer)}")
    print("---- 分类分布 ----")
    for k, v in cat_c.most_common():
        print(f"  {k}: {v}")
    print(f"\nJSON -> {args.out}")


if __name__ == "__main__":
    main()
